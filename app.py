import io
import re
from datetime import date, datetime
from typing import Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st
from openpyxl.styles import PatternFill


st.set_page_config(page_title="FileSort Cleaner", layout="wide")
st.markdown(
    """
    <style>
    [data-testid="stToolbar"] {display: none !important;}
    [data-testid="stDecoration"] {display: none !important;}
    header {display: none !important;}
    #MainMenu {visibility: hidden;}
    </style>
    """,
    unsafe_allow_html=True,
)


CANONICAL_FIELDS = [
    "full_name",
    "first_name",
    "last_name",
    "gender",
    "phone",
    "dob",
    "category",
    "tshirt_size",
    "blood_group",
    "address",
    "city",
    "state",
    "country",
    "emergency_name",
    "emergency_phone",
    "emergency_relation",
]

FIELD_LABELS = {
    "full_name": "Full Name",
    "first_name": "First Name",
    "last_name": "Last Name",
    "gender": "Gender",
    "phone": "Phone",
    "dob": "Date of Birth",
    "category": "Category",
    "tshirt_size": "T-Shirt Size",
    "blood_group": "Blood Group",
    "address": "Address",
    "city": "City",
    "state": "State",
    "country": "Country",
    "emergency_name": "Emergency Contact Name",
    "emergency_phone": "Emergency Contact Phone",
    "emergency_relation": "Emergency Contact Relation",
}

ALIASES = {
    "full_name": ["attendee name", "runner name", "full name", "participant name", "name"],
    "first_name": ["first name", "first_name", "fname"],
    "last_name": ["last name", "last_name", "lname", "surname"],
    "gender": ["gender", "sex"],
    "phone": ["contact number", "mobile", "phone", "telephone", "mobile number", "contact no"],
    "dob": ["date of birth", "dob", "birth date", "date_of_birth"],
    "category": ["ticket_name", "race category", "event category", "distance", "registration category", "category"],
    "tshirt_size": ["t shirt size", "t-shirt size", "tshirt size", "t-shirt", "shirt size"],
    "blood_group": ["blood group", "bloodgroup"],
    "address": ["address"],
    "city": ["city"],
    "state": ["state", "state (india)"],
    "country": ["country"],
    "emergency_name": ["emergency contact name", "emergency name", "emergency contact person"],
    "emergency_phone": ["emergency contact number", "emergency phone", "emergency no", "emergency contact no"],
    "emergency_relation": ["emergency contact relation", "emergency relation", "relationship"],
}

KNOWN_TITLES = {
    "mr",
    "mr.",
    "mrs",
    "mrs.",
    "ms",
    "ms.",
    "dr",
    "dr.",
    "col",
    "col.",
    "colonel",
    "lt",
    "lt.",
    "lt col",
    "capt",
    "capt.",
    "prof",
    "prof.",
}

TSHIRT_MAP = {
    "xxs": "XXS",
    "xs": "XS",
    "s": "S",
    "small": "S",
    "m": "M",
    "medium": "M",
    "l": "L",
    "large": "L",
    "xl": "XL",
    "x-large": "XL",
    "xxl": "XXL",
    "xx-large": "XXL",
    "xxxl": "XXXL",
}

NUM_TO_SIZE = {
    "34": "XXS",
    "36": "XS",
    "38": "S",
    "40": "M",
    "42": "L",
    "44": "XL",
    "46": "XXL",
    "48": "XXXL",
}

VALID_BLOOD_GROUPS = {"A+", "A-", "B+", "B-", "AB+", "AB-", "O+", "O-"}

COUNTRY_MAP = {
    "india": "India",
    "ind": "India",
    "in": "India",
    "usa": "United States",
    "us": "United States",
    "u.s.a.": "United States",
    "uk": "United Kingdom",
    "u.k.": "United Kingdom",
    "uae": "United Arab Emirates",
}

EMERGENCY_RELATION_MAP = {
    "wife": "Wife",
    "husband": "Husband",
    "father": "Father",
    "mother": "Mother",
    "brother": "Brother",
    "sister": "Sister",
    "son": "Son",
    "daughter": "Daughter",
    "spouse": "Spouse",
    "friend": "Friend",
    "colleague": "Colleague",
    "guardian": "Guardian",
    "mentor": "Mentor",
    "uncle": "Uncle",
    "aunt": "Aunt",
}

ERROR_LEGEND = {
    "phone": {"hex": "FF9800", "label": "Phone Error"},
    "dob": {"hex": "FF5252", "label": "Date of Birth Error"},
    "name": {"hex": "448AFF", "label": "Name Issue"},
    "gender": {"hex": "E040FB", "label": "Gender Error"},
    "relation": {"hex": "00E676", "label": "Relation Error"},
    "country": {"hex": "FFEB3B", "label": "Country Error"},
}

COLUMN_FLAG_TO_ERROR = {
    "Phone": {"yellow": "phone"},
    "Emergency Phone": {"yellow": "phone"},
    "Gender": {"yellow": "gender"},
    "Date of Birth": {"red": "dob", "yellow": "dob"},
    "Country": {"yellow": "country"},
    "Emergency Relation": {"yellow": "relation"},
}

NAME_OUTPUT_COLUMNS = ["Full Name", "First Name", "Last Name"]
EMERGENCY_NAME_OUTPUT_COLUMNS = [
    "Emergency Full Name",
    "Emergency First Name",
    "Emergency Last Name",
]


_INVISIBLE_CHARS = re.compile(
    r"[\ufeff\u200b\u200c\u200d\u2060\u00ad"
    r"\u200e\u200f\u202a-\u202e"
    r"\u2061-\u2064\u2066-\u2069"
    r"\ufff9-\ufffb]+"
)
_UNICODE_WHITESPACE = re.compile(r"[\s\xa0\u1680\u180e\u2000-\u200a\u202f\u205f\u3000\v\f]+")


def to_text(value) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    text = str(value).strip()
    if text.lower() == "nan":
        return ""
    return text


def clean_spaces(value) -> str:
    text = to_text(value)
    text = _INVISIBLE_CHARS.sub("", text)
    text = _UNICODE_WHITESPACE.sub(" ", text)
    return text.strip()


def strip_formula(value) -> str:
    text = to_text(value)
    match = re.match(r'^="(.*)"$', text)
    if match:
        text = match.group(1)
    return clean_spaces(text)


def is_blank(value) -> bool:
    raw = strip_formula(value).lower()
    return raw in {"", "na", "n/a", "none", "null", "nan", "-", "0"}


def title_token(token: str) -> str:
    if not token:
        return token
    if "." in token:
        parts = token.split(".")
        fixed = []
        for p in parts:
            if p:
                fixed.append(p[0].upper() + p[1:].lower())
            else:
                fixed.append("")
        return ".".join(fixed)
    return token[0].upper() + token[1:].lower()


def proper_case_text(value: str) -> str:
    value = clean_spaces(value)
    tokens = value.split(" ")
    return " ".join(title_token(token) for token in tokens if token)


def normalize_name_text(raw: str) -> str:
    value = clean_spaces(strip_formula(raw))
    value = re.sub(r"(?i)\b(dr)\.(\S)", r"\1. \2", value)
    return proper_case_text(value)


def split_name(full_name_raw: str) -> Tuple[str, str, str]:
    if is_blank(full_name_raw):
        return ".", ".", "."

    full = normalize_name_text(full_name_raw)
    words = full.split()
    if not words:
        return ".", ".", "."

    first = words[0]
    last_words = words[1:]

    if words[0].lower() in KNOWN_TITLES and len(words) >= 2:
        first = f"{words[0]} {words[1]}"
        last_words = words[2:]

    last = " ".join(last_words) if last_words else "."
    return first, last, full


def merge_name_fields(full_name_raw: str, first_name_raw: str, last_name_raw: str) -> Tuple[str, str, str]:
    """
    Build attendee name using all mapped sources per row.
    - Start from Full Name when available.
    - Override with non-empty First/Last parts when provided.
    - Rebuild Full Name from resolved First + Last.
    """
    full_first, full_last, _ = split_name(full_name_raw)

    first_part = normalize_name_text(first_name_raw) if not is_blank(first_name_raw) else "."
    last_part = normalize_name_text(last_name_raw) if not is_blank(last_name_raw) else "."

    if full_first == "." and first_part == "." and last_part == ".":
        return ".", ".", "."

    first = first_part if first_part != "." else full_first
    last = last_part if last_part != "." else full_last

    # If only one side is present, keep it in first name per single-name rule.
    if first == "." and last != ".":
        first, last = last, "."

    if first == ".":
        return ".", ".", "."

    full = first if last == "." else clean_spaces(f"{first} {last}")
    return first, last, full


def clean_gender(raw: str) -> Tuple[str, Optional[str]]:
    if is_blank(raw):
        return "", None

    value = clean_spaces(strip_formula(raw))
    lower = value.lower()
    lower_plain = re.sub(r"[^a-z]", "", lower)

    male_set = {"male", "m", "man", "maile"}
    female_set = {"female", "f", "woman", "femail", "feamle", "femaile"}
    others_set = {"others", "other", "nonbinary", "nonbinary", "transgender", "prefernottosay"}

    if lower_plain in male_set:
        return "Male", None
    if lower_plain in female_set:
        return "Female", None
    if lower_plain in others_set:
        return "Others", None
    if lower_plain.startswith("m"):
        return "Male", None
    if lower_plain.startswith("f"):
        return "Female", None

    return value, "yellow"


def clean_phone(raw: str) -> Tuple[str, Optional[str]]:
    if is_blank(raw):
        return "", None

    value = clean_spaces(strip_formula(raw))
    digits = re.sub(r"\D", "", value)

    normalized = None
    if len(digits) == 10:
        normalized = digits
    elif len(digits) == 12 and digits.startswith("91"):
        normalized = digits[-10:]
    elif len(digits) == 11 and digits.startswith("0"):
        normalized = digits[-10:]

    if normalized and len(normalized) == 10:
        return normalized, None
    return value, "yellow"


def format_dob_output(dt: pd.Timestamp) -> str:
    """Output as DD-MM-YYYY with leading zeros."""
    return dt.strftime("%d-%m-%Y")


def parse_numeric_date(raw: str) -> Optional[pd.Timestamp]:
    """
    Parse numeric date parts into a timestamp.
    - Ambiguous (both parts <= 12): default DD-MM-YYYY.
    - First part > 12: DD-MM-YYYY.
    - Second part > 12: MM-DD-YYYY.
    - YYYY-MM-DD (ISO) also supported.
    """
    raw = raw.strip().replace(".", "-").replace("/", "-")
    match = re.match(r"^(\d{1,4})-(\d{1,2})-(\d{1,2}|\d{4})$", raw)
    if not match:
        return None

    p1, p2, p3 = int(match.group(1)), int(match.group(2)), int(match.group(3))

    if p1 > 31:
        year, month, day = p1, p2, p3
    elif p1 > 12:
        day, month, year = p1, p2, p3
    elif p2 > 12:
        month, day, year = p1, p2, p3
    else:
        # Ambiguous e.g. 5/12/2005 -> default DD-MM-YYYY (5 Dec 2005)
        day, month, year = p1, p2, p3

    if year < 100:
        year += 2000 if year < 30 else 1900

    try:
        return pd.Timestamp(year=year, month=month, day=day)
    except ValueError:
        return pd.NaT


def clean_dob(raw: str) -> Tuple[str, Optional[str]]:
    if is_blank(raw):
        return "", None

    value = clean_spaces(strip_formula(raw))
    value = value.strip("'\"")

    ordinal_fixed = re.sub(r"(\d+)(st|nd|rd|th)", r"\1", value, flags=re.IGNORECASE)

    # Handle pure numeric Excel serial date.
    if re.fullmatch(r"\d{5}(\.\d+)?", ordinal_fixed):
        serial = float(ordinal_fixed)
        dt = pd.to_datetime(serial, origin="1899-12-30", unit="D", errors="coerce")
    else:
        dt = parse_numeric_date(ordinal_fixed)
        if dt is None:
            dt = pd.to_datetime(ordinal_fixed, dayfirst=True, errors="coerce")

    if pd.isna(dt):
        # impossible numeric date pattern -> red
        if re.match(r"^\d{1,2}[-/.]\d{1,2}[-/.]\d{2,4}$", value):
            return value, "red"
        return value, "yellow"

    if dt.date() > date.today() or dt.year < 1900:
        return value, "red"

    return format_dob_output(dt), None


def clean_tshirt(raw: str) -> Tuple[str, Optional[str]]:
    if is_blank(raw):
        return ".", None

    value = clean_spaces(strip_formula(raw))
    key = value.lower().replace(" ", "")
    key = key.replace("_", "-")

    for word, out in TSHIRT_MAP.items():
        if key == word.replace("-", "") or key == word:
            return out, None

    # Combined formats like XL-44, M40, S-38
    match = re.match(r"^(xxxl|xxl|xl|xxs|xs|s|m|l)[- ]?(\d{2})?$", key)
    if match:
        letter = match.group(1).upper()
        return letter, None

    digits = re.sub(r"\D", "", key)
    if digits in NUM_TO_SIZE:
        return NUM_TO_SIZE[digits], None

    return value, "yellow"


def normalize_blood_group(raw: str) -> Optional[str]:
    """Normalize common blood group text variants to A+/A-/B+/etc."""
    value = clean_spaces(strip_formula(raw)).upper()

    if re.search(r"[/\\_,;|]", value):
        return None

    value = re.sub(r"POSITIVE", "+", value)
    value = re.sub(r"NEGATIVE", "-", value)
    value = re.sub(r"\bPOS\b", "+", value)
    value = re.sub(r"\bNEG\b", "-", value)
    value = value.replace("+VE", "+").replace("-VE", "-")
    value = re.sub(r"\bRH\b", "", value)

    if re.search(r"\d", value):
        return None

    compact = re.sub(r"[^ABO+\-]", "", value)
    compact = re.sub(r"(AB|A|B|O)-\+", r"\1+", compact)
    compact = re.sub(r"(AB|A|B|O)--", r"\1-", compact)

    if compact in VALID_BLOOD_GROUPS:
        return compact
    return None


def clean_blood_group(raw: str) -> Tuple[str, Optional[str]]:
    if is_blank(raw):
        return ".", None

    original = clean_spaces(strip_formula(raw))
    normalized = normalize_blood_group(original)
    if normalized:
        return normalized, None

    return original, None


def clean_address_like(raw: str) -> Tuple[str, Optional[str]]:
    if is_blank(raw):
        return ".", None
    value = proper_case_text(strip_formula(raw))
    return value, None


def clean_country(raw: str) -> Tuple[str, Optional[str]]:
    if is_blank(raw):
        return "India", None
    value = proper_case_text(strip_formula(raw))
    key = value.lower()
    if key in COUNTRY_MAP:
        return COUNTRY_MAP[key], None
    if re.search(r"[a-zA-Z]", value):
        return value, None
    return value, "yellow"


def clean_emergency_relation(raw: str) -> Tuple[str, Optional[str]]:
    if is_blank(raw):
        return ".", None
    value = proper_case_text(strip_formula(raw))
    key = value.lower()
    if key in EMERGENCY_RELATION_MAP:
        return EMERGENCY_RELATION_MAP[key], None
    if key == "bhai":
        return "Brother", None
    return value, "yellow"


def is_dob_column(col_name: str) -> bool:
    norm = re.sub(r"[^a-z0-9]+", " ", str(col_name).lower()).strip()
    return norm.startswith("date of birth") or norm == "dob" or norm == "date of birth"


def find_dob_source_columns(df: pd.DataFrame, primary: Optional[str]) -> List[str]:
    """Primary mapped DOB column plus any duplicate DOB columns in the file."""
    if not primary:
        return []
    cols = []
    for col in df.columns:
        if col == primary or is_dob_column(col):
            cols.append(col)
    if primary in cols:
        cols.remove(primary)
        return [primary] + cols
    return [primary]


def first_non_blank_value(row: pd.Series, columns: List[str]) -> str:
    for col in columns:
        value = row[col]
        if not is_blank(value):
            return value
    return ""


def parse_dob_for_excel(value: str) -> Optional[datetime]:
    """Convert cleaned DD-MM-YYYY text to datetime for Excel sorting."""
    if is_blank(value) or value == ".":
        return None
    match = re.match(r"^(\d{2})-(\d{2})-(\d{4})$", str(value).strip())
    if not match:
        return None
    day, month, year = int(match.group(1)), int(match.group(2)), int(match.group(3))
    try:
        return datetime(year, month, day)
    except ValueError:
        return None


def calculate_age_from_dob(dob_value: str, age_as_on: date) -> str:
    """Calculate completed age (years) from cleaned DOB using a manual as-on date."""
    parsed = parse_dob_for_excel(dob_value)
    if not parsed:
        return ""
    dob_date = parsed.date()
    years = age_as_on.year - dob_date.year
    if (age_as_on.month, age_as_on.day) < (dob_date.month, dob_date.day):
        years -= 1
    return str(max(0, years))


def guess_mapping(columns) -> Dict[str, Optional[str]]:
    guessed = {field: None for field in CANONICAL_FIELDS}
    lowered = {col: re.sub(r"[^a-z0-9]+", " ", str(col).lower()).strip() for col in columns}

    for field, names in ALIASES.items():
        for col, normalized in lowered.items():
            for alias in names:
                alias_norm = re.sub(r"[^a-z0-9]+", " ", alias.lower()).strip()
                if normalized == alias_norm or alias_norm in normalized:
                    if guessed[field] is None:
                        guessed[field] = col
                    break
            if guessed[field] is not None:
                break
    return guessed


def read_uploaded_file(uploaded_file) -> pd.DataFrame:
    filename = uploaded_file.name.lower()
    if filename.endswith(".csv"):
        return pd.read_csv(uploaded_file, dtype=str, keep_default_na=False).fillna("")
    if filename.endswith(".xlsx"):
        return pd.read_excel(uploaded_file, dtype=str).fillna("")
    if filename.endswith(".xls"):
        # Try old Excel first, fallback to HTML-table style xls.
        try:
            return pd.read_excel(uploaded_file, dtype=str).fillna("")
        except Exception:
            uploaded_file.seek(0)
            html = uploaded_file.read().decode("utf-8", errors="replace")
            df = pd.read_html(io.StringIO(html), header=0)[0]
            if len(df) > 0 and str(df.iloc[0, 0]).strip().lower() in {"first name", "name"}:
                df.columns = df.iloc[0]
                df = df.iloc[1:].reset_index(drop=True)
            return df.astype(str).fillna("")
    raise ValueError("Unsupported file format")


def mark_name_issue(cell_flags: Dict[Tuple[int, str], str], row_idx: int, columns: List[str]):
    for column in columns:
        cell_flags[(row_idx, column)] = "name"


def render_error_legend():
    st.markdown("**Error Color Legend**")
    legend_cols = st.columns(len(ERROR_LEGEND))
    for col, (_, info) in zip(legend_cols, ERROR_LEGEND.items()):
        with col:
            st.markdown(
                f'<div style="background-color:#{info["hex"]}; padding:12px 8px; border-radius:6px; '
                f'border:2px solid #222; text-align:center; color:#111; font-weight:700;">'
                f'{info["label"]}</div>',
                unsafe_allow_html=True,
            )


def count_errors_by_type(cell_flags: Dict[Tuple[int, str], str]) -> Dict[str, int]:
    counts: Dict[str, int] = {}
    for error_type in cell_flags.values():
        counts[error_type] = counts.get(error_type, 0) + 1
    return counts


def style_preview_dataframe(df: pd.DataFrame, cell_flags: Dict[Tuple[int, str], str]):
    def row_style(row: pd.Series):
        styles = []
        for col_name in row.index:
            error_type = cell_flags.get((row.name, col_name))
            if error_type and error_type in ERROR_LEGEND:
                color = ERROR_LEGEND[error_type]["hex"]
                styles.append(f"background-color: #{color}")
            else:
                styles.append("")
        return styles

    return df.style.apply(row_style, axis=1)


def apply_cleaning(
    df: pd.DataFrame,
    mapping: Dict[str, Optional[str]],
    category_map: Dict[str, str],
    tshirt_map: Dict[str, str],
    age_as_on: date,
) -> Tuple[pd.DataFrame, Dict[Tuple[int, str], str]]:
    out = df.copy()
    cell_flags: Dict[Tuple[int, str], str] = {}

    def mark_cell(row_idx: int, column: str, flag: Optional[str]):
        if not flag:
            return
        error_type = COLUMN_FLAG_TO_ERROR.get(column, {}).get(flag)
        if error_type:
            cell_flags[(row_idx, column)] = error_type

    # Name handling: if full name exists use it to derive first/last/full by your rules.
    full_name_col = mapping.get("full_name")
    first_name_col = mapping.get("first_name")
    last_name_col = mapping.get("last_name")

    if full_name_col or first_name_col or last_name_col:
        cleaned_first = []
        cleaned_last = []
        cleaned_full = []

        for i in range(len(out)):
            full_raw = out.at[i, full_name_col] if full_name_col else ""
            first_raw = out.at[i, first_name_col] if first_name_col else ""
            last_raw = out.at[i, last_name_col] if last_name_col else ""
            first, last, full = merge_name_fields(full_raw, first_raw, last_raw)
            cleaned_first.append(first)
            cleaned_last.append(last)
            cleaned_full.append(full)
            if first == "." and last == "." and full == ".":
                mark_name_issue(cell_flags, i, NAME_OUTPUT_COLUMNS)

        out["First Name"] = cleaned_first
        out["Last Name"] = cleaned_last
        out["Full Name"] = cleaned_full

    for target, cleaner in [
        ("Gender", clean_gender),
        ("Phone", clean_phone),
        ("Blood Group", clean_blood_group),
        ("Address", clean_address_like),
        ("City", clean_address_like),
        ("State", clean_address_like),
        ("Emergency Phone", clean_phone),
        ("Emergency Relation", clean_emergency_relation),
    ]:
        field_key = {
            "Gender": "gender",
            "Phone": "phone",
            "Blood Group": "blood_group",
            "Address": "address",
            "City": "city",
            "State": "state",
            "Emergency Phone": "emergency_phone",
            "Emergency Relation": "emergency_relation",
        }[target]
        src_col = mapping.get(field_key)
        if src_col:
            cleaned_values = []
            for i in range(len(out)):
                raw = out.at[i, src_col]
                cleaned, flag = cleaner(raw)
                cleaned_values.append(cleaned)
                mark_cell(i, target, flag)
            out[target] = cleaned_values

    # DOB: merge first non-empty across duplicate DOB columns.
    dob_primary = mapping.get("dob")
    dob_cols = find_dob_source_columns(out, dob_primary)
    if dob_cols:
        dob_values = []
        age_values = []
        for i in range(len(out)):
            raw = first_non_blank_value(out.iloc[i], dob_cols)
            cleaned, flag = clean_dob(raw)
            dob_values.append(cleaned)
            age_values.append(calculate_age_from_dob(cleaned, age_as_on))
            mark_cell(i, "Date of Birth", flag)
        out["Date of Birth"] = dob_values
        out["Age"] = age_values

    # Emergency name: same first/last/full rules as attendee name.
    emergency_name_col = mapping.get("emergency_name")
    if emergency_name_col:
        em_first, em_last, em_full = [], [], []
        for i in range(len(out)):
            first, last, full = split_name(out.at[i, emergency_name_col])
            em_first.append(first)
            em_last.append(last)
            em_full.append(full)
            if first == "." and last == "." and full == ".":
                mark_name_issue(cell_flags, i, EMERGENCY_NAME_OUTPUT_COLUMNS)
        out["Emergency First Name"] = em_first
        out["Emergency Last Name"] = em_last
        out["Emergency Full Name"] = em_full

    # Category mapping
    cat_col = mapping.get("category")
    if cat_col:
        category_values = []
        for i in range(len(out)):
            raw = out.at[i, cat_col]
            if is_blank(raw):
                category_values.append(".")
                continue
            raw_clean = clean_spaces(strip_formula(raw))
            mapped = category_map.get(raw_clean, raw_clean)
            category_values.append(mapped)
        out["Category"] = category_values

    # T-shirt mapping (manual mapping only, same behavior as category mapping)
    tshirt_col = mapping.get("tshirt_size")
    if tshirt_col:
        tshirt_values = []
        for i in range(len(out)):
            raw = out.at[i, tshirt_col]
            if is_blank(raw):
                tshirt_values.append(".")
                continue
            raw_clean = clean_spaces(strip_formula(raw))
            mapped = tshirt_map.get(raw_clean, raw_clean)
            tshirt_values.append(mapped)
        out["T-Shirt Size"] = tshirt_values

    # Country logic
    country_col = mapping.get("country")
    if country_col:
        values = []
        for i in range(len(out)):
            cleaned, flag = clean_country(out.at[i, country_col])
            values.append(cleaned)
            mark_cell(i, "Country", flag)
        out["Country"] = values
    else:
        out["Country"] = ["India"] * len(out)

    # Keep cleaned-first order.
    cleaned_order = [
        "Full Name",
        "First Name",
        "Last Name",
        "Gender",
        "Phone",
        "Date of Birth",
        "Age",
        "Category",
        "T-Shirt Size",
        "Blood Group",
        "Address",
        "City",
        "State",
        "Country",
        "Emergency First Name",
        "Emergency Last Name",
        "Emergency Full Name",
        "Emergency Phone",
        "Emergency Relation",
    ]
    cleaned_order = [c for c in cleaned_order if c in out.columns]

    mapped_sources = {src for src in mapping.values() if src}
    if dob_primary:
        mapped_sources.update(find_dob_source_columns(out, dob_primary))

    extra_cols = [
        c for c in df.columns if c not in cleaned_order and c not in mapped_sources
    ]
    final_order = cleaned_order + extra_cols
    final = out[final_order].copy()
    for col in cleaned_order:
        final[col] = [clean_spaces(v) for v in final[col]]
    return sort_cleaned_df(final, cell_flags)


GENDER_SORT_ORDER = {"Male": 0, "Female": 1, "Others": 2}


def sort_cleaned_df(
    df: pd.DataFrame, cell_flags: Dict[Tuple[int, str], str]
) -> Tuple[pd.DataFrame, Dict[Tuple[int, str], str]]:
    """Sort by category, then gender (Male, Female, Others)."""
    if df.empty:
        return df, cell_flags

    sorted_df = df.copy()
    sorted_df["_orig_idx"] = range(len(sorted_df))

    sort_cols: List[str] = []
    ascending: List[bool] = []

    if "Category" in sorted_df.columns:
        sort_cols.append("Category")
        ascending.append(True)

    if "Gender" in sorted_df.columns:
        sorted_df["_gender_order"] = sorted_df["Gender"].map(lambda g: GENDER_SORT_ORDER.get(g, 3))
        sort_cols.append("_gender_order")
        ascending.append(True)

    if sort_cols:
        sorted_df = sorted_df.sort_values(by=sort_cols, ascending=ascending, kind="stable")

    orig_indices = sorted_df["_orig_idx"].astype(int).tolist()
    drop_cols = [c for c in ["_orig_idx", "_gender_order"] if c in sorted_df.columns]
    sorted_df = sorted_df.drop(columns=drop_cols).reset_index(drop=True)

    new_flags: Dict[Tuple[int, str], str] = {}
    for new_idx, orig_idx in enumerate(orig_indices):
        for col_name in df.columns:
            key = (orig_idx, col_name)
            if key in cell_flags:
                new_flags[(new_idx, col_name)] = cell_flags[key]
    return sorted_df, new_flags


def df_to_excel_with_highlight(df: pd.DataFrame, cell_flags: Dict[Tuple[int, str], str]) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Cleaned")
        ws = writer.book["Cleaned"]

        col_name_to_idx = {name: idx + 1 for idx, name in enumerate(df.columns)}
        error_fills = {
            error_type: PatternFill(
                start_color=info["hex"], end_color=info["hex"], fill_type="solid"
            )
            for error_type, info in ERROR_LEGEND.items()
        }

        dob_col_idx = col_name_to_idx.get("Date of Birth")
        phone_col_idx = col_name_to_idx.get("Phone")
        emergency_phone_col_idx = col_name_to_idx.get("Emergency Phone")
        age_col_idx = col_name_to_idx.get("Age")

        for (row_idx, col_name), error_type in cell_flags.items():
            excel_row = row_idx + 2
            col_idx = col_name_to_idx.get(col_name)
            fill = error_fills.get(error_type)
            if col_idx and fill:
                ws.cell(row=excel_row, column=col_idx).fill = fill

        # Write valid DOB values as real Excel dates for proper sorting.
        if dob_col_idx:
            for row_idx in range(len(df)):
                excel_row = row_idx + 2
                raw_val = df.iloc[row_idx]["Date of Birth"]
                parsed = parse_dob_for_excel(raw_val)
                if parsed:
                    cell = ws.cell(row=excel_row, column=dob_col_idx)
                    cell.value = parsed
                    cell.number_format = "DD-MM-YYYY"

        # Write valid phone values as numbers to avoid Excel text warnings.
        for col_idx, col_name in [
            (phone_col_idx, "Phone"),
            (emergency_phone_col_idx, "Emergency Phone"),
        ]:
            if not col_idx or col_name not in df.columns:
                continue
            for row_idx in range(len(df)):
                excel_row = row_idx + 2
                raw_val = clean_spaces(df.iloc[row_idx][col_name])
                if re.fullmatch(r"\d{10}", raw_val):
                    cell = ws.cell(row=excel_row, column=col_idx)
                    cell.value = int(raw_val)
                    cell.number_format = "0"

        # Write age as numeric cells to avoid Excel text warnings.
        if age_col_idx and "Age" in df.columns:
            for row_idx in range(len(df)):
                excel_row = row_idx + 2
                raw_val = clean_spaces(df.iloc[row_idx]["Age"])
                if re.fullmatch(r"\d+", raw_val):
                    cell = ws.cell(row=excel_row, column=age_col_idx)
                    cell.value = int(raw_val)
                    cell.number_format = "0"
    output.seek(0)
    return output.read()


st.title("FileSort Cleaner")
st.caption("Upload -> scan -> manual map -> category map -> clean -> download")

uploaded = st.file_uploader("Upload CSV/XLSX/XLS", type=["csv", "xlsx", "xls"])

if uploaded:
    try:
        src_df = read_uploaded_file(uploaded)
    except Exception as exc:
        st.error(f"Could not read file: {exc}")
        st.stop()

    st.success(f"Loaded file with {len(src_df)} rows and {len(src_df.columns)} columns.")
    st.dataframe(src_df.head(8), use_container_width=True)

    guessed = guess_mapping(src_df.columns)
    options = ["<None>"] + list(src_df.columns)

    st.subheader("Step 1: Manual Column Mapping (after scan)")
    st.write("Confirm the correct source column for each target field.")

    mapping: Dict[str, Optional[str]] = {}
    cols_ui = st.columns(2)
    for idx, field in enumerate(CANONICAL_FIELDS):
        col_ui = cols_ui[idx % 2]
        default = guessed.get(field)
        default_index = options.index(default) if default in options else 0
        picked = col_ui.selectbox(
            FIELD_LABELS[field],
            options=options,
            index=default_index,
            key=f"map_{field}",
        )
        mapping[field] = None if picked == "<None>" else picked

    dob_col = mapping.get("dob")
    if dob_col:
        extra_dob_cols = [c for c in find_dob_source_columns(src_df, dob_col) if c != dob_col]
        if extra_dob_cols:
            st.caption(
                "DOB merge enabled: will use first non-empty value from -> "
                + ", ".join([dob_col] + extra_dob_cols)
            )
    age_as_on = st.date_input("Age calculation date (as on)", value=date.today())

    category_map: Dict[str, str] = {}
    tshirt_map: Dict[str, str] = {}
    cat_col = mapping.get("category")
    if cat_col:
        st.subheader("Step 2: Category Manual Mapping")
        raw_categories = (
            src_df[cat_col]
            .fillna("")
            .astype(str)
            .map(strip_formula)
            .map(clean_spaces)
            .replace("", ".")
            .drop_duplicates()
            .tolist()
        )
        raw_categories = sorted(raw_categories)
        cat_counts = (
            src_df[cat_col]
            .fillna("")
            .astype(str)
            .map(strip_formula)
            .map(clean_spaces)
            .replace("", ".")
            .value_counts()
            .to_dict()
        )
        if "category_map_df" not in st.session_state or st.session_state.get("category_map_source") != tuple(raw_categories):
            st.session_state["category_map_df"] = pd.DataFrame(
                {"Raw Category": raw_categories, "Mapped Category": raw_categories}
            )
            st.session_state["category_map_source"] = tuple(raw_categories)
            st.session_state["category_groups"] = [{"selected": [], "name": ""}]
        elif "category_groups" not in st.session_state:
            st.session_state["category_groups"] = [{"selected": [], "name": ""}]

        st.info("Map similar categories together in groups, then apply all mappings.")

        # Multi-group mapping UI.
        st.markdown("**Quick Group Mapping**")
        for idx, grp in enumerate(st.session_state["category_groups"]):
            st.markdown(f"**Group {idx + 1}**")
            used_by_other_groups = set()
            for j, other_grp in enumerate(st.session_state["category_groups"]):
                if j == idx:
                    continue
                used_by_other_groups.update(other_grp.get("selected", []))
            available_options = [
                cat for cat in raw_categories if cat not in used_by_other_groups or cat in grp.get("selected", [])
            ]
            current_default = [cat for cat in grp.get("selected", []) if cat in available_options]
            left, right = st.columns([2, 1])
            with left:
                grp["selected"] = st.multiselect(
                    f"Select raw categories (Group {idx + 1})",
                    options=available_options,
                    default=current_default,
                    key=f"bulk_raw_select_{idx}",
                    placeholder="Pick one or more categories",
                )
            with right:
                grp["name"] = st.text_input(
                    f"Mapped name (Group {idx + 1})",
                    value=grp.get("name", ""),
                    key=f"bulk_mapped_name_{idx}",
                    placeholder="e.g. 10KM Defense",
                )


        btn_col1, btn_col2, btn_col3 = st.columns([1, 1, 1])
        with btn_col1:
            add_group_clicked = st.button("Add One More Group", use_container_width=True)
        with btn_col2:
            apply_clicked = st.button("Apply All Group Mappings", use_container_width=True)
        with btn_col3:
            reset_clicked = st.button("Reset All Category Mappings", use_container_width=True)

        if add_group_clicked:
            st.session_state["category_groups"].append({"selected": [], "name": ""})
            st.rerun()

        if apply_clicked:
            tmp_df = st.session_state["category_map_df"].copy()
            applied = 0
            for grp in st.session_state["category_groups"]:
                mapped_name = clean_spaces(grp.get("name", ""))
                selected_raw = grp.get("selected", [])
                if not selected_raw or not mapped_name:
                    continue
                tmp_df.loc[tmp_df["Raw Category"].isin(selected_raw), "Mapped Category"] = mapped_name
                applied += len(selected_raw)
            if applied == 0:
                st.warning("Add at least one valid group (selected categories + mapped name).")
            else:
                st.session_state["category_map_df"] = tmp_df
                st.success(f"Applied group mappings to {applied} category selections.")

        if reset_clicked:
            st.session_state["category_map_df"] = pd.DataFrame(
                {"Raw Category": raw_categories, "Mapped Category": raw_categories}
            )
            st.session_state["category_groups"] = [{"selected": [], "name": ""}]
            st.success("Reset complete. All mapped values restored to original raw categories.")

        map_df = st.session_state["category_map_df"]

        # Friendly summary table.
        summary_df = map_df.copy()
        summary_df["Rows"] = summary_df["Raw Category"].map(lambda x: cat_counts.get(x, 0))
        summary_df = summary_df[["Raw Category", "Rows", "Mapped Category"]].sort_values(
            by=["Mapped Category", "Raw Category"]
        )
        st.markdown("**Current Mapping Summary**")
        st.dataframe(summary_df, hide_index=True, use_container_width=True)

        for _, row in map_df.iterrows():
            category_map[str(row["Raw Category"])] = str(row["Mapped Category"]).strip() or "."
    else:
        st.info("No category column selected. Category output will be '.'")

    tshirt_col = mapping.get("tshirt_size")
    if tshirt_col:
        st.subheader("Step 3: T-Shirt Size Manual Mapping")
        raw_tshirts = (
            src_df[tshirt_col]
            .fillna("")
            .astype(str)
            .map(strip_formula)
            .map(clean_spaces)
            .replace("", ".")
            .drop_duplicates()
            .tolist()
        )
        raw_tshirts = sorted(raw_tshirts)
        tshirt_counts = (
            src_df[tshirt_col]
            .fillna("")
            .astype(str)
            .map(strip_formula)
            .map(clean_spaces)
            .replace("", ".")
            .value_counts()
            .to_dict()
        )
        if "tshirt_map_df" not in st.session_state or st.session_state.get("tshirt_map_source") != tuple(raw_tshirts):
            st.session_state["tshirt_map_df"] = pd.DataFrame(
                {"Raw T-Shirt Size": raw_tshirts, "Mapped T-Shirt Size": raw_tshirts}
            )
            st.session_state["tshirt_map_source"] = tuple(raw_tshirts)
            st.session_state["tshirt_groups"] = [{"selected": [], "name": ""}]
        elif "tshirt_groups" not in st.session_state:
            st.session_state["tshirt_groups"] = [{"selected": [], "name": ""}]

        st.info("Map similar T-shirt values together in groups, then apply all mappings.")
        st.markdown("**Quick T-Shirt Group Mapping**")
        for idx, grp in enumerate(st.session_state["tshirt_groups"]):
            st.markdown(f"**T-Shirt Group {idx + 1}**")
            used_by_other_groups = set()
            for j, other_grp in enumerate(st.session_state["tshirt_groups"]):
                if j == idx:
                    continue
                used_by_other_groups.update(other_grp.get("selected", []))
            available_options = [
                size for size in raw_tshirts if size not in used_by_other_groups or size in grp.get("selected", [])
            ]
            current_default = [size for size in grp.get("selected", []) if size in available_options]
            left, right = st.columns([2, 1])
            with left:
                grp["selected"] = st.multiselect(
                    f"Select raw T-shirt sizes (Group {idx + 1})",
                    options=available_options,
                    default=current_default,
                    key=f"tshirt_bulk_raw_select_{idx}",
                    placeholder="Pick one or more T-shirt values",
                )
            with right:
                grp["name"] = st.text_input(
                    f"Mapped T-shirt size (Group {idx + 1})",
                    value=grp.get("name", ""),
                    key=f"tshirt_bulk_mapped_name_{idx}",
                    placeholder="e.g. M",
                )

        tshirt_btn_col1, tshirt_btn_col2, tshirt_btn_col3 = st.columns([1, 1, 1])
        with tshirt_btn_col1:
            tshirt_add_group_clicked = st.button("Add One More T-Shirt Group", use_container_width=True)
        with tshirt_btn_col2:
            tshirt_apply_clicked = st.button("Apply All T-Shirt Group Mappings", use_container_width=True)
        with tshirt_btn_col3:
            tshirt_reset_clicked = st.button("Reset All T-Shirt Mappings", use_container_width=True)

        if tshirt_add_group_clicked:
            st.session_state["tshirt_groups"].append({"selected": [], "name": ""})
            st.rerun()

        if tshirt_apply_clicked:
            tmp_df = st.session_state["tshirt_map_df"].copy()
            applied = 0
            for grp in st.session_state["tshirt_groups"]:
                mapped_name = clean_spaces(grp.get("name", ""))
                selected_raw = grp.get("selected", [])
                if not selected_raw or not mapped_name:
                    continue
                tmp_df.loc[tmp_df["Raw T-Shirt Size"].isin(selected_raw), "Mapped T-Shirt Size"] = mapped_name
                applied += len(selected_raw)
            if applied == 0:
                st.warning("Add at least one valid T-shirt group (selected values + mapped name).")
            else:
                st.session_state["tshirt_map_df"] = tmp_df
                st.success(f"Applied T-shirt mappings to {applied} selections.")

        if tshirt_reset_clicked:
            st.session_state["tshirt_map_df"] = pd.DataFrame(
                {"Raw T-Shirt Size": raw_tshirts, "Mapped T-Shirt Size": raw_tshirts}
            )
            st.session_state["tshirt_groups"] = [{"selected": [], "name": ""}]
            st.success("Reset complete. All T-shirt mapped values restored to original raw values.")

        tshirt_df = st.session_state["tshirt_map_df"]
        tshirt_summary_df = tshirt_df.copy()
        tshirt_summary_df["Rows"] = tshirt_summary_df["Raw T-Shirt Size"].map(lambda x: tshirt_counts.get(x, 0))
        tshirt_summary_df = tshirt_summary_df[
            ["Raw T-Shirt Size", "Rows", "Mapped T-Shirt Size"]
        ].sort_values(by=["Mapped T-Shirt Size", "Raw T-Shirt Size"])
        st.markdown("**Current T-Shirt Mapping Summary**")
        st.dataframe(tshirt_summary_df, hide_index=True, use_container_width=True)

        for _, row in tshirt_df.iterrows():
            tshirt_map[str(row["Raw T-Shirt Size"])] = str(row["Mapped T-Shirt Size"]).strip() or "."
    else:
        st.info("No T-shirt size column selected. T-Shirt Size output will be '.'")

    st.subheader("Error Highlight Guide")
    render_error_legend()

    if st.button("Run Cleaning", type="primary"):
        cleaned_df, cell_flags = apply_cleaning(src_df, mapping, category_map, tshirt_map, age_as_on)
        error_counts = count_errors_by_type(cell_flags)
        if error_counts:
            summary_parts = [
                f"{ERROR_LEGEND[error_type]['label']}: {count}"
                for error_type, count in sorted(error_counts.items())
                if error_type in ERROR_LEGEND
            ]
            st.success(f"Cleaning completed. Flagged cells -> {' | '.join(summary_parts)}")
        else:
            st.success("Cleaning completed. No flagged errors found.")

        render_error_legend()
        preview_df = cleaned_df.head(20)
        if cell_flags:
            st.dataframe(style_preview_dataframe(preview_df, cell_flags), use_container_width=True)
        else:
            st.dataframe(preview_df, use_container_width=True)

        excel_bytes = df_to_excel_with_highlight(cleaned_df, cell_flags)
        st.download_button(
            label="Download Cleaned Excel",
            data=excel_bytes,
            file_name="cleaned_output.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

